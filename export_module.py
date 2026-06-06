import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime, date
from PIL import Image, ImageDraw, ImageFont
import database as db

def _default_export_dir():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if os.path.isdir(desktop):
        return os.path.join(desktop, "du_lieu_output")
    return os.path.join(os.path.dirname(__file__), "exports")


EXPORT_DIR = _default_export_dir()
os.makedirs(EXPORT_DIR, exist_ok=True)


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

STATUS_VI = {
    "present": "Có mặt",
    "absent": "Vắng",
    "excused": "Vắng có phép",
    "half": "Học 1/2",
    "scanning": "Đang quét",
}

STATUS_COLOR = {
    "present": "C6EFCE",  # xanh lá nhạt
    "absent": "FFC7CE",   # đỏ nhạt
    "excused": "D9EAD3",  # xanh xám nhạt
    "half": "FFEB9C",     # vàng nhạt
    "scanning": "BDD7EE", # xanh dương nhạt
}


def export_by_date(target_date: str) -> str:
    """Xuất điểm danh theo ngày ra file Excel"""
    settings = db.get_all_settings()
    class_name = settings.get("class_name", "Lớp học")
    total_periods = int(settings.get("total_periods", 3))

    students = db.get_all_students()
    if not students:
        return None, "Chưa có sinh viên nào"

    sessions = db.get_sessions_by_date(target_date)
    # Map session_id by period
    session_map = {s["period"]: s for s in sessions}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Điểm danh {target_date}"

    # ── Header Block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"📋 BẢNG ĐIỂM DANH - {class_name.upper()}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = _header_fill("1A3C5E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Ngày: {_format_date_vi(target_date)}   |   Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}"
    ws["A2"].font = Font(name="Calibri", size=11, color="FFFFFF", italic=True)
    ws["A2"].fill = _header_fill("2E6DA4")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Column Headers ────────────────────────────────────────────────────────
    headers = ["STT", "Mã SV", "Họ và Tên"]
    for p in range(1, total_periods + 1):
        s = session_map.get(p)
        time_str = ""
        if s:
            time_str = f"\n({s['start_time']}–{s['end_time']})"
        headers.append(f"Tiết {p}{time_str}")
    headers += ["Có mặt", "Vắng", "Học 1/2", "Tỉ lệ (%)"]

    col_count = len(headers)
    # Extend merge if needed
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    ws.merge_cells(f"A2:{get_column_letter(col_count)}2")

    row3 = ws.row_dimensions[3]
    row3.height = 40
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = _header_fill("34495E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    # ── Data Rows ─────────────────────────────────────────────────────────────
    for row_idx, student in enumerate(students, start=1):
        sid = student["student_id"]
        data_row = [row_idx, sid, student["name"]]
        present_cnt = 0
        absent_cnt = 0
        half_cnt = 0

        for p in range(1, total_periods + 1):
            sess = session_map.get(p)
            if sess:
                att_list = db.get_attendance_by_session(sess["id"])
                att = next((a for a in att_list if a["student_id"] == sid), None)
                status = att["status"] if att else "absent"
            else:
                status = "absent"
            data_row.append(STATUS_VI.get(status, "-"))
            if status == "present":
                present_cnt += 1
            elif status == "absent":
                absent_cnt += 1
            elif status == "half":
                half_cnt += 1

        total = present_cnt + absent_cnt + half_cnt + (0.5 * half_cnt)
        ratio = round((present_cnt + half_cnt * 0.5) / total_periods * 100, 1) if total_periods > 0 else 0
        data_row += [present_cnt, absent_cnt, half_cnt, f"{ratio}%"]

        excel_row = row_idx + 3
        ws.row_dimensions[excel_row].height = 20
        fill_color = "F0F4F8" if row_idx % 2 == 0 else "FFFFFF"

        for col_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border("thin")

            # Màu ô trạng thái tiết học
            col_is_period = 3 < col_idx <= 3 + total_periods
            if col_is_period:
                for stat_key, stat_vi in STATUS_VI.items():
                    if val == stat_vi:
                        cell.fill = PatternFill("solid", fgColor=STATUS_COLOR[stat_key])
                        break
            elif col_idx > 3 + total_periods:
                cell.fill = PatternFill("solid", fgColor=fill_color)
            else:
                cell.fill = PatternFill("solid", fgColor=fill_color)

            # Tỉ lệ màu
            if col_idx == col_count:
                try:
                    pct = float(str(val).replace("%", ""))
                    if pct >= 80:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif pct >= 50:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C5700")
                    else:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                except:
                    pass

    # ── Summary Row ───────────────────────────────────────────────────────────
    summary_row = len(students) + 4
    ws.row_dimensions[summary_row].height = 22
    ws.cell(summary_row, 1, "TỔNG").font = Font(bold=True, color="FFFFFF")
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"].fill = _header_fill("1A3C5E")
    ws[f"A{summary_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{summary_row}"].font = Font(name="Calibri", bold=True, color="FFFFFF")

    for p_col in range(4, 4 + total_periods):
        col_data = [ws.cell(r, p_col).value for r in range(4, 4 + len(students))]
        present_total = col_data.count("Có mặt")
        ws.cell(summary_row, p_col, f"{present_total}/{len(students)}").font = Font(bold=True)
        ws.cell(summary_row, p_col).alignment = Alignment(horizontal="center")
        ws.cell(summary_row, p_col).fill = _header_fill("BDD7EE")
        ws.cell(summary_row, p_col).border = _border()

    # ── Column Widths ─────────────────────────────────────────────────────────
    col_widths = [6, 14, 28] + [14] * total_periods + [10, 8, 10, 10]
    for i, w in enumerate(col_widths, start=1):
        if i <= col_count:
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Chart Sheet ───────────────────────────────────────────────────────────
    _add_summary_chart(wb, students, sessions, total_periods, target_date)

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = summary_row + 2
    ws.cell(legend_row, 1, "CHÚ THÍCH:").font = Font(bold=True, size=10)
    for i, (stat, color) in enumerate(STATUS_COLOR.items()):
        r = legend_row + i
        ws.cell(r, 1, STATUS_VI[stat])
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(r, 1).border = _border()
        ws.cell(r, 1).alignment = Alignment(horizontal="center")

    filename = f"diemdanh_{target_date}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath, f"Xuất thành công: {filename}"


def _add_summary_chart(wb, students, sessions, total_periods, target_date):
    """Thêm sheet biểu đồ thống kê"""
    ws2 = wb.create_sheet("📊 Thống kê")
    ws2["A1"] = "Sinh viên"
    ws2["B1"] = "Có mặt"
    ws2["C1"] = "Vắng"
    ws2["D1"] = "Học 1/2"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2["C1"].font = Font(bold=True)
    ws2["D1"].font = Font(bold=True)

    session_map = {s["period"]: s for s in sessions}

    for row_idx, student in enumerate(students, start=2):
        sid = student["student_id"]
        ws2.cell(row_idx, 1, student["name"])
        p = a = h = 0
        for period in range(1, total_periods + 1):
            sess = session_map.get(period)
            if sess:
                att_list = db.get_attendance_by_session(sess["id"])
                att = next((x for x in att_list if x["student_id"] == sid), None)
                status = att["status"] if att else "absent"
                if status == "present": p += 1
                elif status == "absent": a += 1
                elif status == "half": h += 1
        ws2.cell(row_idx, 2, p)
        ws2.cell(row_idx, 3, a)
        ws2.cell(row_idx, 4, h)

    if len(students) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Thống kê điểm danh {target_date}"
        chart.y_axis.title = "Số tiết"
        chart.x_axis.title = "Sinh viên"
        chart.style = 10
        chart.width = 20
        chart.height = 14

        data = Reference(ws2, min_col=2, max_col=4, min_row=1, max_row=len(students)+1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(students)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "2ECC71"
        chart.series[1].graphicalProperties.solidFill = "E74C3C"
        if len(chart.series) > 2:
            chart.series[2].graphicalProperties.solidFill = "F39C12"

        ws2.add_chart(chart, "F2")


def export_monthly_report(year: int, month: int) -> str:
    """Xuất báo cáo tháng"""
    stats = db.get_monthly_stats(year, month)
    settings = db.get_all_settings()
    class_name = settings.get("class_name", "Lớp học")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Bao cao T{month}-{year}"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"BÁO CÁO THÁNG {month}/{year} - {class_name.upper()}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _header_fill("1A3C5E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["STT", "Mã SV", "Họ và Tên", "Có mặt", "Vắng", "Học 1/2", "Tỉ lệ (%)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, h)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.fill = _header_fill("34495E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    for col in "DEFG":
        ws.column_dimensions[col].width = 12

    for row_idx, s in enumerate(stats, 1):
        total = s["total_sessions"] if s["total_sessions"] else 1
        ratio = round((s["present_count"] + s["half_count"] * 0.5) / total * 100, 1)
        row_data = [row_idx, s["student_id"], s["name"],
                    s["present_count"], s["absent_count"], s["half_count"], f"{ratio}%"]
        fill = "F0F4F8" if row_idx % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx + 2, col, val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
            cell.fill = PatternFill("solid", fgColor=fill)

    filename = f"baocao_thang_{month:02d}_{year}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath, f"Xuất thành công: {filename}"


def _load_pdf_font(size=24, bold=False):
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeuib.ttf" if bold else r"C:\Windows\Fonts\segoeui.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def _fit_text(draw, text, font, max_width):
    text = "" if text is None else str(text)
    if draw.textlength(text, font=font) <= max_width:
        return text
    ellipsis = "..."
    while text and draw.textlength(text + ellipsis, font=font) > max_width:
        text = text[:-1]
    return text + ellipsis if text else ellipsis


def _draw_pdf_table(filepath, title, subtitle, headers, rows):
    page_w, page_h = 1754, 1240
    margin = 60
    title_font = _load_pdf_font(32, True)
    sub_font = _load_pdf_font(18)
    header_font = _load_pdf_font(17, True)
    cell_font = _load_pdf_font(16)
    row_h = 44
    header_h = 52
    top_h = 126
    usable_h = page_h - margin - top_h - 54
    rows_per_page = max(1, (usable_h - header_h) // row_h)

    if not rows:
        rows = [["Không có dữ liệu"] + [""] * (len(headers) - 1)]

    pages = []
    total_pages = (len(rows) + rows_per_page - 1) // rows_per_page
    for page_start in range(0, len(rows), rows_per_page):
        chunk = rows[page_start:page_start + rows_per_page]
        img = Image.new("RGB", (page_w, page_h), "white")
        draw = ImageDraw.Draw(img)

        draw.rectangle([0, 0, page_w, 90], fill="#1A3C5E")
        draw.text((margin, 22), title, fill="white", font=title_font)
        draw.text((margin, 98), subtitle, fill="#34495E", font=sub_font)

        table_x = margin
        table_y = margin + top_h
        table_w = page_w - margin * 2
        fixed = [70, 160]
        remaining = table_w - sum(fixed)
        if len(headers) <= 2:
            col_widths = [table_w // len(headers)] * len(headers)
        else:
            name_w = min(360, max(250, int(remaining * 0.28)))
            other_count = max(1, len(headers) - 3)
            other_w = max(92, int((remaining - name_w) / other_count))
            col_widths = fixed + [name_w] + [other_w] * other_count
            col_widths[-1] += table_w - sum(col_widths)

        x = table_x
        for col, header in enumerate(headers):
            w = col_widths[col]
            draw.rectangle([x, table_y, x + w, table_y + header_h], fill="#34495E", outline="#2A3F55")
            draw.text((x + 7, table_y + 15), _fit_text(draw, header, header_font, w - 14), fill="white", font=header_font)
            x += w

        y = table_y + header_h
        fills = {"Có mặt": "#C6EFCE", "Vắng": "#FFC7CE", "Học 1/2": "#FFEB9C", "Đang quét": "#BDD7EE"}
        for ridx, row in enumerate(chunk):
            x = table_x
            base_fill = "#F5F8FB" if ridx % 2 else "white"
            for col, value in enumerate(row):
                w = col_widths[col]
                fill = fills.get(str(value), base_fill)
                draw.rectangle([x, y, x + w, y + row_h], fill=fill, outline="#D7DEE8")
                draw.text((x + 7, y + 12), _fit_text(draw, value, cell_font, w - 14), fill="#1B2530", font=cell_font)
                x += w
            y += row_h

        page_no = len(pages) + 1
        draw.text((page_w - margin - 160, page_h - 36), f"Trang {page_no}/{total_pages}", fill="#637083", font=sub_font)
        pages.append(img)

    pages[0].save(filepath, "PDF", resolution=150.0, save_all=True, append_images=pages[1:])


def export_by_date_pdf(target_date: str):
    settings = db.get_all_settings()
    class_name = settings.get("class_name", "Lớp học")
    total_periods = int(settings.get("total_periods", 3))
    students = db.get_all_students()
    if not students:
        return None, "Chưa có sinh viên nào"

    sessions = db.get_sessions_by_date(target_date)
    session_map = {s["period"]: s for s in sessions}
    headers = ["STT", "Mã SV", "Họ và Tên"] + [f"Tiết {p}" for p in range(1, total_periods + 1)] + ["Tỉ lệ"]
    rows = []
    for idx, student in enumerate(students, 1):
        sid = student["student_id"]
        row = [idx, sid, student["name"]]
        score = 0
        for p in range(1, total_periods + 1):
            sess = session_map.get(p)
            status = "absent"
            if sess:
                att_list = db.get_attendance_by_session(sess["id"])
                att = next((a for a in att_list if a["student_id"] == sid), None)
                status = att["status"] if att else "absent"
            row.append(STATUS_VI.get(status, "-"))
            if status == "present":
                score += 1
            elif status == "half":
                score += 0.5
        ratio = round(score / total_periods * 100, 1) if total_periods else 0
        row.append(f"{ratio}%")
        rows.append(row)

    filename = f"diemdanh_{target_date}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)
    _draw_pdf_table(
        filepath,
        f"BẢNG ĐIỂM DANH - {class_name.upper()}",
        f"Ngày: {_format_date_vi(target_date)} | Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}",
        headers,
        rows,
    )
    return filepath, f"Xuất thành công: {filename}"


def export_monthly_report_pdf(year: int, month: int):
    stats = db.get_monthly_stats(year, month)
    settings = db.get_all_settings()
    class_name = settings.get("class_name", "Lớp học")
    headers = ["STT", "Mã SV", "Họ và Tên", "Có mặt", "Vắng", "Học 1/2", "Tỉ lệ"]
    rows = []
    for idx, s in enumerate(stats, 1):
        total = s["total_sessions"] if s["total_sessions"] else 1
        ratio = round((s["present_count"] + s["half_count"] * 0.5) / total * 100, 1)
        rows.append([idx, s["student_id"], s["name"], s["present_count"], s["absent_count"], s["half_count"], f"{ratio}%"])

    filename = f"baocao_thang_{month:02d}_{year}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)
    _draw_pdf_table(
        filepath,
        f"BÁO CÁO THÁNG {month}/{year} - {class_name.upper()}",
        f"Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}",
        headers,
        rows,
    )
    return filepath, f"Xuất thành công: {filename}"


def _format_date_vi(date_str):
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        days_vi = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
        return f"{days_vi[d.weekday()]}, {d.day:02d}/{d.month:02d}/{d.year}"
    except:
        return date_str
